import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 난수 시드 설정
np.random.seed(42)

# 데이터 생성 함수들
def generate_patient_data(n=100):
    """환자 기본 정보"""
    patient_ids = [f"PAT{str(i).zfill(4)}" for i in range(1, n+1)]
    ages = np.random.randint(35, 85, n)
    genders = np.random.choice(['M', 'F'], n)

    # 순환기계 질환
    diseases = ['고혈압', '협심증', '심근경색', '부정맥', '심부전', '고지혈증', '말초동맥질환']
    primary_disease = np.random.choice(diseases, n)

    comorbidities = []
    for _ in range(n):
        comorb = np.random.choice(['당뇨병', '만성신질환', '뇌졸중', '없음'],
                                   p=[0.3, 0.2, 0.15, 0.35])
        comorbidities.append(comorb)

    enrollment_dates = [datetime(2023, 1, 1) + timedelta(days=int(x))
                       for x in np.random.uniform(0, 365, n)]

    return pd.DataFrame({
        '환자ID': patient_ids,
        '나이': ages,
        '성별': genders,
        '주진단': primary_disease,
        '동반질환': comorbidities,
        '등록일': [d.strftime('%Y-%m-%d') for d in enrollment_dates],
        '체중(kg)': np.random.uniform(50, 100, n).round(1),
        '키(cm)': np.random.uniform(150, 190, n).round(1)
    })

def generate_treatment_data(patient_data):
    """투약 데이터"""
    drugs = ['아텐롤', '암로디핀', '리시노프릴', '심바스타틴', '아스피린',
             '와파린', '메토프롤올', '에나라프릴', '로수바스타틴', '프로프라놀올']

    dosages = ['5mg', '10mg', '20mg', '50mg', '100mg', '75mg', '150mg']
    frequencies = ['1일 1회', '1일 2회', '1일 3회']

    treatment_list = []
    for idx, row in patient_data.iterrows():
        num_drugs = np.random.randint(1, 4)
        selected_drugs = np.random.choice(drugs, num_drugs, replace=False)

        for drug in selected_drugs:
            start_date = datetime.strptime(row['등록일'], '%Y-%m-%d') + timedelta(days=np.random.randint(0, 7))
            treatment_list.append({
                '환자ID': row['환자ID'],
                '약물명': drug,
                '용량': np.random.choice(dosages),
                '복용빈도': np.random.choice(frequencies),
                '투약시작일': start_date.strftime('%Y-%m-%d'),
                '투약종료일': (start_date + timedelta(days=np.random.randint(30, 365))).strftime('%Y-%m-%d')
                              if np.random.random() > 0.3 else '',
                '순응도(%)': np.random.choice([85, 90, 95, 100])
            })

    return pd.DataFrame(treatment_list)

def generate_lab_results(patient_data):
    """검사 결과 데이터"""
    lab_list = []
    test_dates = pd.date_range(start='2023-01-15', end='2023-12-31', freq='3M')

    for idx, row in patient_data.iterrows():
        for test_date in test_dates:
            age = row['나이']
            lab_list.append({
                '환자ID': row['환자ID'],
                '검사일': test_date.strftime('%Y-%m-%d'),
                '수축기혈압(mmHg)': int(np.random.normal(135, 15)),
                '이완기혈압(mmHg)': int(np.random.normal(85, 10)),
                '심박수(bpm)': int(np.random.normal(70, 10)),
                '콜레스테롤(mg/dL)': int(np.random.normal(200, 50)),
                'LDL(mg/dL)': int(np.random.normal(130, 40)),
                'HDL(mg/dL)': int(np.random.normal(40, 15)),
                'AST(U/L)': int(np.random.normal(35, 15)),
                'ALT(U/L)': int(np.random.normal(30, 15)),
                'Creatinine(mg/dL)': round(np.random.normal(0.9, 0.3), 2)
            })

    return pd.DataFrame(lab_list)

def generate_adverse_events(patient_data):
    """이상반응 데이터"""
    adverse_events = ['두통', '어지러움', '피로', '기침', '변비', '구역질',
                     '불면증', '근육통', '발진', '저혈당증']
    severity_levels = ['경미', '중등도', '중증']

    ae_list = []
    for idx, row in patient_data.iterrows():
        if np.random.random() > 0.6:  # 40%의 환자만 이상반응 보고
            num_events = np.random.randint(1, 3)
            for _ in range(num_events):
                event_date = datetime.strptime(row['등록일'], '%Y-%m-%d') + timedelta(days=np.random.randint(1, 180))
                ae_list.append({
                    '환자ID': row['환자ID'],
                    '이상반응명': np.random.choice(adverse_events),
                    '심각도': np.random.choice(severity_levels, p=[0.6, 0.3, 0.1]),
                    '발생일': event_date.strftime('%Y-%m-%d'),
                    '인과관계': np.random.choice(['관련있음', '관련불명확', '관련없음'], p=[0.5, 0.3, 0.2]),
                    '조치': np.random.choice(['없음', '용량감소', '약물중단', '약물변경'], p=[0.4, 0.3, 0.2, 0.1]),
                    '결과': np.random.choice(['회복', '지속', '악화'], p=[0.7, 0.2, 0.1])
                })

    return pd.DataFrame(ae_list)

def generate_efficacy_data(patient_data):
    """약효 평가 데이터"""
    efficacy_list = []

    for idx, row in patient_data.iterrows():
        eval_dates = [datetime.strptime(row['등록일'], '%Y-%m-%d') + timedelta(days=x)
                      for x in [0, 30, 60, 90, 180]]

        for week, eval_date in enumerate(eval_dates):
            improvement = max(0, 100 - (week * 15 + np.random.normal(0, 5)))
            efficacy_list.append({
                '환자ID': row['환자ID'],
                '평가일': eval_date.strftime('%Y-%m-%d'),
                '주차': week * 4,
                '증상개선율(%)': int(improvement),
                'LVEF(%)': int(np.random.normal(55 + week*3, 5)),
                '6분보행거리(m)': int(np.random.normal(400 + week*20, 30))
            })

    return pd.DataFrame(efficacy_list)

# 데이터 생성
print("데이터 생성 중...")
patient_df = generate_patient_data(100)
treatment_df = generate_treatment_data(patient_df)
lab_df = generate_lab_results(patient_df)
ae_df = generate_adverse_events(patient_df)
efficacy_df = generate_efficacy_data(patient_df)

print(f"✓ 환자 정보: {len(patient_df)}명")
print(f"✓ 투약 데이터: {len(treatment_df)}건")
print(f"✓ 검사 결과: {len(lab_df)}건")
print(f"✓ 이상반응: {len(ae_df)}건")
print(f"✓ 약효 평가: {len(efficacy_df)}건")

# Excel 파일 생성
output_file = 'C:\\workspace\\순환기계_임상데이터.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    patient_df.to_excel(writer, sheet_name='환자정보', index=False)
    treatment_df.to_excel(writer, sheet_name='투약데이터', index=False)
    lab_df.to_excel(writer, sheet_name='검사결과', index=False)
    ae_df.to_excel(writer, sheet_name='이상반응', index=False)
    efficacy_df.to_excel(writer, sheet_name='약효평가', index=False)

    # 스타일링
    wb = writer.book
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws.columns:
            max_length = 0
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 20)

print(f"\n✓ Excel 파일 생성 완료!")
print(f"  저장 위치: {output_file}")
